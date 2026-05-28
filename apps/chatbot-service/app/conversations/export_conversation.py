"""Service for exporting conversations to Excel."""
from __future__ import annotations

import uuid
from datetime import datetime
from io import BytesIO
from typing import Optional

from domain.db_service.message_services import GettingAllMessagesForExportInput
from domain.db_service.message_services import GettingAllMessagesForExportService
from joint.base import BaseModel
from joint.base import BaseService
from joint.logging import get_logger
from joint.settings.settings import PostgresSettings
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

logger = get_logger(__name__)


class ConversationExportInput(BaseModel):
    """Input model for exporting conversations."""

    user_id: uuid.UUID
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None


class ConversationExportOutput(BaseModel):
    """Output model for exporting conversations."""

    status: bool
    message: str = ''
    file_content: Optional[bytes] = None
    filename: str = ''

    class Config:
        """Pydantic config."""

        arbitrary_types_allowed = True


class ConversationExportService(BaseService):
    """App service for exporting conversations to Excel."""

    postgres_settings: PostgresSettings

    async def process(
        self,
        inputs: ConversationExportInput,
        db_session=None,
    ) -> ConversationExportOutput:
        """Export conversations to Excel file.

        Args:
            inputs: Export input with user_id and optional date filters.
            db_session: Optional database session.

        Returns:
            Output with Excel file content as bytes.
        """
        try:
            logger.info(f"Exporting conversations for user: {inputs.user_id}")

            # Get all messages from domain service
            domain_input = GettingAllMessagesForExportInput(
                user_id=inputs.user_id,
                start_date=inputs.start_date,
                end_date=inputs.end_date,
            )

            domain_service = GettingAllMessagesForExportService(
                settings=self.postgres_settings,
            )
            result = domain_service.process(domain_input, db_session)

            if not result.status:
                return ConversationExportOutput(
                    status=False,
                    message=result.message,
                )

            # Build Excel file
            file_content = self._build_excel(result.data)

            # Generate filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"conversations_export_{timestamp}.xlsx"

            logger.info(f"Export completed: {len(result.data)} messages")
            return ConversationExportOutput(
                status=True,
                message=f"Exported {len(result.data)} messages",
                file_content=file_content,
                filename=filename,
            )

        except Exception as e:
            logger.error(f"Error exporting conversations: {e}", exc_info=True)
            return ConversationExportOutput(
                status=False,
                message=f"Export failed: {e}",
            )

    def _build_excel(self, messages: list) -> bytes:
        """Build Excel file from messages.

        Args:
            messages: List of ExportMessageItem.

        Returns:
            Excel file content as bytes.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'Q&A Export'

        # Define headers
        headers = [
            'Timestamp', 'User Question',
            'Bot Answer', 'Conversation Title',
        ]
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment

        # Group messages by pairs (user + assistant)
        row = 2
        current_question = None
        current_title = None
        current_timestamp = None

        for msg in messages:
            if msg.message_type == 'human':
                current_question = msg.contents
                current_title = msg.conversation_title
                current_timestamp = msg.timestamp
            elif msg.message_type == 'ai' and current_question is not None:
                # Write Q&A pair
                timestamp_str = current_timestamp.strftime(
                    '%Y-%m-%d %H:%M:%S',
                ) if current_timestamp else ''
                ws.cell(row=row, column=1, value=timestamp_str)
                ws.cell(row=row, column=2, value=current_question)
                ws.cell(row=row, column=3, value=msg.contents)
                ws.cell(row=row, column=4, value=current_title or 'Untitled')
                row += 1
                current_question = None

        # Auto-adjust column widths
        for col in range(1, 5):
            max_length = max(
                len(str(ws.cell(row=r, column=col).value or ''))
                for r in range(1, row)
            ) if row > 1 else 10
            ws.column_dimensions[
                get_column_letter(
                    col,
                )
            ].width = min(max_length + 2, 50)

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
